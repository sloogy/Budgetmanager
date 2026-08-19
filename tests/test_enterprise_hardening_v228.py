from __future__ import annotations

import base64
import hashlib
import json
import zipfile
from pathlib import Path

import pytest
from cryptography.hazmat.primitives.asymmetric.ed25519 import Ed25519PrivateKey
from cryptography.hazmat.primitives.serialization import (
    Encoding,
    PrivateFormat,
    NoEncryption,
)
from openpyxl import Workbook

from model.restore_bundle import (
    LegacyBundleIntegrityError,
    read_manifest,
    upgrade_legacy_bundle,
    verify_bundle,
)
from tools.generate_sbom import generate_sbom
from tools.verify_hashed_lock import validate
from updater.manifest_signing import (
    ManifestSignatureError,
    public_key_base64,
    sign_manifest_bytes,
    verify_manifest_signature,
)
from utils.secure_excel import (
    UnsafeExcelFileError,
    load_workbook_safely,
    validate_excel_archive,
)

ROOT = Path(__file__).resolve().parents[1]


def _private_key_b64(key: Ed25519PrivateKey) -> str:
    raw = key.private_bytes(Encoding.Raw, PrivateFormat.Raw, NoEncryption())
    return base64.b64encode(raw).decode("ascii")


def test_manifest_signature_accepts_valid_and_rejects_tampering(monkeypatch):
    key = Ed25519PrivateKey.generate()
    manifest = b'{"version":"2.2.28"}\n'
    signature = sign_manifest_bytes(manifest, key)
    monkeypatch.setenv("BUDGETMANAGER_UPDATE_PUBLIC_KEY_B64", public_key_base64(key))
    verify_manifest_signature(manifest, signature)
    with pytest.raises(ManifestSignatureError):
        verify_manifest_signature(manifest + b" ", signature)
    with pytest.raises(ManifestSignatureError):
        verify_manifest_signature(manifest, b"not-base64")


def test_private_and_public_update_keys_must_match(tmp_path):
    from updater.manifest_signing import sign_manifest_file

    manifest = tmp_path / "latest.json"
    manifest.write_text('{"version":"2.2.28"}\n', encoding="utf-8")
    key = Ed25519PrivateKey.generate()
    wrong = Ed25519PrivateKey.generate()
    with pytest.raises(ManifestSignatureError):
        sign_manifest_file(
            manifest,
            private_key_b64=_private_key_b64(key),
            expected_public_key_b64=public_key_base64(wrong),
        )


def test_excel_valid_workbook_is_loaded(tmp_path):
    path = tmp_path / "categories.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.append(["Konto", "Kategorie"])
    ws.append(["Ausgaben", "Miete"])
    wb.save(path)
    assert validate_excel_archive(path) == path.resolve()
    loaded = load_workbook_safely(path, read_only=True, data_only=True)
    try:
        assert loaded.active["B2"].value == "Miete"
    finally:
        loaded.close()


def test_excel_rejects_path_traversal_and_dtd(tmp_path):
    traversal = tmp_path / "traversal.xlsx"
    with zipfile.ZipFile(traversal, "w") as archive:
        archive.writestr("../evil.xml", "<x/>")
    with pytest.raises(UnsafeExcelFileError, match="Unsicherer Pfad"):
        validate_excel_archive(traversal)

    dtd = tmp_path / "dtd.xlsx"
    with zipfile.ZipFile(dtd, "w") as archive:
        archive.writestr(
            "[Content_Types].xml", '<!DOCTYPE x [<!ENTITY e "x">]><x>&e;</x>'
        )
    with pytest.raises(UnsafeExcelFileError, match="DTD/ENTITY"):
        validate_excel_archive(dtd)


def test_excel_rejects_extreme_compression_ratio(tmp_path):
    bomb = tmp_path / "bomb.xlsx"
    with zipfile.ZipFile(bomb, "w", compression=zipfile.ZIP_DEFLATED) as archive:
        archive.writestr("xl/worksheets/sheet1.xml", b"0" * (2 * 1024 * 1024))
    with pytest.raises(UnsafeExcelFileError, match="Kompressionsrate"):
        validate_excel_archive(bomb)


def test_legacy_backup_is_rejected_then_upgraded(tmp_path):
    source = tmp_path / "legacy.bmr"
    db_bytes = b"SQLite format 3\x00" + b"test" * 50
    manifest = {
        "created_at": "2026-07-18T00:00:00",
        "app": "Budgetmanager",
        "app_version": "2.0.0",
        "db_file": "database.db",
        "encryption": "db",
    }
    with zipfile.ZipFile(source, "w", compression=zipfile.ZIP_DEFLATED) as archive:
        archive.writestr("manifest.json", json.dumps(manifest))
        archive.writestr("database.db", db_bytes)

    with pytest.raises(LegacyBundleIntegrityError):
        verify_bundle(source)
    assert verify_bundle(source, allow_legacy_without_hash=True) == "database.db"

    upgraded = upgrade_legacy_bundle(source, tmp_path / "legacy_verified.bmr")
    assert source.exists()
    assert verify_bundle(upgraded) == "database.db"
    upgraded_manifest = read_manifest(upgraded)
    assert upgraded_manifest["sha256"] == hashlib.sha256(db_bytes).hexdigest()
    assert upgraded_manifest["integrity_format"] == "sha256-v1"


def test_all_release_lockfiles_are_exact_and_hashed():
    pairs = [
        ("requirements.lock", "requirements.in"),
        ("requirements-dev.lock", "requirements-dev.in"),
        ("requirements-build.lock", "requirements-build.in"),
    ]
    for lock, direct in pairs:
        assert validate(ROOT / lock, ROOT / direct) == []


def test_sbom_contains_runtime_hashes(tmp_path):
    out = generate_sbom(
        ROOT / "requirements.lock", tmp_path / "sbom.json", app_version="2.2.28"
    )
    data = json.loads(out.read_text(encoding="utf-8"))
    assert data["bomFormat"] == "CycloneDX"
    assert data["metadata"]["component"]["version"] == "2.2.28"
    assert data["components"]
    assert all(component["hashes"] for component in data["components"])


def test_release_pipeline_contains_the_required_core_build_contract():
    workflow = (ROOT / ".github/workflows/build.yml").read_text(encoding="utf-8")
    required = [
        "python tools/verify_hashed_lock.py",
        "python -m pip install --require-hashes -r requirements-build.lock",
        "python -m pip install --require-hashes -r requirements-dev.lock",
        "python -m black --check model/",
        "python -m mypy model/",
        "python -m pytest tests/ -v -ra --tb=short",
        "pyinstaller BudgetManager.spec --noconfirm --clean",
        "Build Windows installer",
        "tools/build_release_assets.py",
        "SHA256SUMS.txt",
        "softprops/action-gh-release@v2",
    ]
    for marker in required:
        assert marker in workflow


def test_crypto_documentation_does_not_claim_aes256_fernet():
    user_model = (ROOT / "model/user_model.py").read_text(encoding="utf-8")
    assert "AES-256" not in user_model
    assert "AES-128-CBC + HMAC-SHA256" in user_model


def test_manifest_key_loading_candidates_and_file_signing(monkeypatch, tmp_path):
    import updater.manifest_signing as signing

    key = Ed25519PrivateKey.generate()
    private_b64 = _private_key_b64(key)
    public_b64 = public_key_base64(key)

    # String-Decodierung, Längenprüfung und fail-closed ohne Vertrauensanker.
    assert signing.private_key_from_base64(private_b64)
    with pytest.raises(ManifestSignatureError, match="32 Bytes"):
        signing.private_key_from_base64(base64.b64encode(b"short").decode("ascii"))

    monkeypatch.delenv("BUDGETMANAGER_UPDATE_PUBLIC_KEY_B64", raising=False)
    monkeypatch.setattr(signing, "public_key_candidates", lambda: ())
    with pytest.raises(ManifestSignatureError, match="Kein eingebetteter"):
        signing.load_trusted_public_key()

    # Eingebetteter Schlüssel wird aus der ersten vorhandenen Kandidatendatei geladen.
    key_file = tmp_path / signing.PUBLIC_KEY_FILENAME
    key_file.write_text(public_b64 + "\n", encoding="ascii")
    monkeypatch.setattr(
        signing, "public_key_candidates", lambda: (tmp_path / "missing", key_file)
    )
    loaded = signing.load_trusted_public_key()
    manifest = b'{"version":"2.2.28"}\n'
    signature = signing.sign_manifest_bytes(manifest, key)
    signing.verify_manifest_signature(manifest, signature, public_key=loaded)

    manifest_path = tmp_path / "latest.json"
    manifest_path.write_bytes(manifest)
    out = signing.sign_manifest_file(
        manifest_path,
        private_key_b64=private_b64,
        expected_public_key_b64=public_b64,
    )
    assert out == tmp_path / "latest.json.sig"
    signing.verify_manifest_signature(manifest, out.read_bytes(), public_key=loaded)

    custom = tmp_path / "detached.signature"
    assert (
        signing.sign_manifest_file(
            manifest_path,
            private_key_b64=private_b64,
            signature_path=custom,
        )
        == custom
    )


def test_manifest_public_key_candidates_deduplicate_meipass(monkeypatch, tmp_path):
    import updater.manifest_signing as signing

    monkeypatch.setattr(signing.sys, "_MEIPASS", str(tmp_path), raising=False)
    monkeypatch.setattr(signing.sys, "executable", str(tmp_path / "BudgetManager"))
    candidates = signing.public_key_candidates()
    assert len(candidates) == len(set(candidates))
    assert tmp_path / "resources" / signing.PUBLIC_KEY_FILENAME in candidates


def test_excel_rejects_invalid_container_shapes(monkeypatch, tmp_path):
    import stat
    import utils.secure_excel as secure_excel

    with pytest.raises(FileNotFoundError):
        validate_excel_archive(tmp_path / "missing.xlsx")

    wrong_suffix = tmp_path / "categories.zip"
    wrong_suffix.write_bytes(b"not-an-xlsx")
    with pytest.raises(UnsafeExcelFileError, match="Nur XLSX"):
        validate_excel_archive(wrong_suffix)

    empty = tmp_path / "empty.xlsx"
    empty.write_bytes(b"")
    with pytest.raises(UnsafeExcelFileError, match="leer oder grösser"):
        validate_excel_archive(empty)

    invalid = tmp_path / "invalid.xlsx"
    invalid.write_bytes(b"not-a-zip")
    with pytest.raises(UnsafeExcelFileError, match="kein gültiger"):
        validate_excel_archive(invalid)

    no_members = tmp_path / "no-members.xlsx"
    with zipfile.ZipFile(no_members, "w"):
        pass
    with pytest.raises(UnsafeExcelFileError, match="zu viele Einträge"):
        validate_excel_archive(no_members)

    symlink = tmp_path / "symlink.xlsx"
    info = zipfile.ZipInfo("xl/link.xml")
    info.external_attr = (stat.S_IFLNK | 0o777) << 16
    with zipfile.ZipFile(symlink, "w") as archive:
        archive.writestr(info, b"target")
    with pytest.raises(UnsafeExcelFileError, match="Symlink"):
        validate_excel_archive(symlink)

    oversized_member = tmp_path / "oversized-member.xlsx"
    with zipfile.ZipFile(oversized_member, "w") as archive:
        archive.writestr("xl/data.bin", b"12")
    monkeypatch.setattr(secure_excel, "MAX_EXCEL_MEMBER_BYTES", 1)
    with pytest.raises(UnsafeExcelFileError, match="Eintrag.*zu gross"):
        secure_excel.validate_excel_archive(oversized_member)

    monkeypatch.setattr(secure_excel, "MAX_EXCEL_MEMBER_BYTES", 1024)
    monkeypatch.setattr(secure_excel, "MAX_EXCEL_UNCOMPRESSED_BYTES", 1)
    with pytest.raises(UnsafeExcelFileError, match="entpackt.*gross"):
        secure_excel.validate_excel_archive(oversized_member)
