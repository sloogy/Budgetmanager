from __future__ import annotations

import sys
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.utils import range_boundaries

from model.database import open_db
from model.app_paths import db_path as portable_db_path
from model.migrations import migrate_all
from model.category_model import CategoryModel
from model.budget_model import BudgetModel

MONTH_COLS = list(range(5, 17))  # E..P -> 12 Monate

def _read_budget_block(ws, start_row:int, end_row:int, year:int, typ:str, budget: BudgetModel):
    # In Budgett Pannung: Kategorie in Spalte C, Monatswerte in E..P
    for r in range(start_row, end_row+1):
        cat = ws.cell(row=r, column=3).value
        if not cat or not isinstance(cat, str):
            continue
        for month, col in enumerate(MONTH_COLS, start=1):
            v = ws.cell(row=r, column=col).value
            if v is None or v == "":
                continue
            try:
                amt=float(v)
            except Exception:
                continue
            budget.set_amount(year, month, typ, cat.strip(), amt)

def main() -> int:
    if len(sys.argv) < 2:
        print("Usage: python tools/import_excel.py <Budgetplan.xlsm>")
        return 2
    xlsm = Path(sys.argv[1]).expanduser()
    if not xlsm.exists():
        print("File not found:", xlsm)
        return 2

    # Portable: nutze immer ./data/budgetmanager.db
    dbp = portable_db_path()
    dbp.parent.mkdir(parents=True, exist_ok=True)
    conn = open_db(str(dbp))
    migrate_all(conn)
    cats = CategoryModel(conn)
    budget = BudgetModel(conn)

    wb = load_workbook(xlsm, data_only=True, keep_vba=True)
    ws = wb["Budgett Pannung"]

    # Jahr 1 ist in E5 (start_jahr) -> data_only kann Wert geben, sonst fallback 2024
    y1 = ws["E5"].value if isinstance(ws["E5"].value, (int,float)) else 2024
    y1 = int(y1)
    y2 = y1 + 1

    # Kategorien aus Tabellen
    t_eink = ws._tables["Einkommen"].ref
    t_ausg = ws._tables["Ausgaben"].ref
    t_ersp = ws._tables["Ersparnisse"].ref

    def read_table_names(tref, typ):
        minc, minr, maxc, maxr = range_boundaries(tref)
        for r in range(minr+1, maxr+1):
            name=ws.cell(row=r, column=minc).value
            if isinstance(name, str) and name.strip():
                # Fix-Flag: bei Ausgaben ist in Excel Flag in Spalte B
                is_fix=False
                if typ=="Ausgaben":
                    flag=ws.cell(row=r, column=2).value
                    is_fix = bool(flag and str(flag).strip())
                cats.upsert(typ, name.strip(), is_fix, True if is_fix else False)

    read_table_names(t_eink, "Einkommen")
    read_table_names(t_ausg, "Ausgaben")
    read_table_names(t_ersp, "Ersparnisse")

    # Budget-Blöcke (aus Analyse):
    # Einkommen: rows 10-12
    _read_budget_block(ws, 10, 12, y1, "Einkommen", budget)
    _read_budget_block(ws, 10, 12, y2, "Einkommen", budget)   # year2 values in S..AB nicht gelesen (optional)

    # Ausgaben: rows 23-43
    _read_budget_block(ws, 23, 43, y1, "Ausgaben", budget)

    # Ersparnisse: rows 57-60
    _read_budget_block(ws, 57, 60, y1, "Ersparnisse", budget)

    print("Import fertig. Jahr:", y1)
    return 0

if __name__ == "__main__":
    raise SystemExit(main())
