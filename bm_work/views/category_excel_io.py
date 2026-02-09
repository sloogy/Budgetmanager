from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
import re
import sqlite3
from typing import Any

from openpyxl import Workbook, load_workbook


TYP_ALIASES = {
    "einnahmen": "Einkommen",
    "income": "Einkommen",
    "einkommen": "Einkommen",
    "lohn": "Einkommen",
    "ausgaben": "Ausgaben",
    "expenses": "Ausgaben",
    "ersparnisse": "Ersparnisse",
    "sparen": "Ersparnisse",
    "savings": "Ersparnisse",
}


def _norm_typ(value: Any) -> str | None:
    if value is None:
        return None
    s = str(value).strip()
    if not s:
        return None
    key = s.strip().lower()
    return TYP_ALIASES.get(key, s)


def _as_bool(value: Any) -> bool:
    if value is None:
        return False
    if isinstance(value, bool):
        return value
    s = str(value).strip().lower()
    return s in {"1", "true", "ja", "j", "yes", "y", "x", "✓", "ok"}


def _as_int_day(value: Any, default: int = 1) -> int:
    try:
        d = int(value)
    except Exception:
        d = int(default)
    if d < 1:
        return 1
    if d > 31:
        return 31
    return d


def _split_path(path: str) -> list[str]:
    # Unterstützt: "Wohnen › Miete", "Wohnen > Miete", "Wohnen / Miete"
    parts = re.split(r"\s*(?:›|»|>|/|\\)\s*", str(path).strip())
    return [p.strip() for p in parts if p and str(p).strip()]


@dataclass
class CategoryImportResult:
    inserted: int
    updated: int
    skipped: int
    warnings: list[str]


def export_category_template_xlsx(out_path: Path) -> Path:
    """Erstellt eine einfache Excel-Vorlage für Kategorien."""
    out_path = Path(out_path)
    if out_path.suffix.lower() != ".xlsx":
        out_path = out_path.with_suffix(".xlsx")

    wb = Workbook()

    ws = wb.active
    ws.title = "Kategorien"

    headers = [
        "Typ",
        "Pfad",
        "Fix (0/1)",
        "Wiederkehrend (0/1)",
        "Tag (1-31)",
    ]
    ws.append(headers)

    # Beispiele (kann der User löschen)
    ws.append(["Ausgaben", "Wohnen › Miete", 1, 1, 1])
    ws.append(["Ausgaben", "Gesundheit › Krankenkasse › Prämie", 1, 1, 1])
    ws.append(["Einkommen", "Lohn", 0, 1, 25])
    ws.append(["Ersparnisse", "Notgroschen", 0, 1, 1])

    # Spaltenbreite
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 44
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 10

    # Info-Sheet
    info = wb.create_sheet("Info")
    info["A1"] = "Budgetmanager – Kategorien-Import"
    info["A3"] = "Spalten:"
    info["A4"] = "Typ: Einkommen / Ausgaben / Ersparnisse (Synonyme: Einnahmen, Sparen)"
    info["A5"] = "Pfad: z.B. Gesundheit › Krankenkasse › Prämie"
    info["A6"] = "Fix: 1 = Fixkosten (⭐)"
    info["A7"] = "Wiederkehrend: 1 = wiederkehrend (∞)"
    info["A8"] = "Tag: Fälligkeitstag (1–31)"
    info["A10"] = "Hinweis: Du kannst die Beispielzeilen löschen und deine Struktur eintragen."

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return out_path


def import_categories_from_xlsx(conn: sqlite3.Connection, xlsx_path: Path) -> CategoryImportResult:
    """Importiert Kategorien (inkl. Baum-Pfad) aus einer Excel-Datei."""
    xlsx_path = Path(xlsx_path)
    if not xlsx_path.exists():
        raise FileNotFoundError(str(xlsx_path))

    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb["Kategorien"] if "Kategorien" in wb.sheetnames else wb.active

    # Header lesen
    header_row = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    header_map: dict[str, int] = {}
    for idx, h in enumerate(header_row):
        if h is None:
            continue
        key = str(h).strip().lower()
        header_map[key] = idx

    def col(*names: str) -> int | None:
        for n in names:
            if n in header_map:
                return header_map[n]
        return None

    c_typ = col("typ")
    c_path = col("pfad", "path")
    c_fix = col("fix (0/1)", "fix", "fixkosten")
    c_rec = col("wiederkehrend (0/1)", "wiederkehrend", "recurring")
    c_day = col("tag (1-31)", "tag", "day")

    if c_typ is None or c_path is None:
        raise ValueError("Excel-Header muss mindestens 'Typ' und 'Pfad' enthalten (Sheet: Kategorien).")

    # Prüfen ob parent_id Spalte existiert
    cols = {r[1] for r in conn.execute("PRAGMA table_info(categories)").fetchall()}
    has_parent = "parent_id" in cols

    warnings: list[str] = []
    inserted = 0
    updated = 0
    skipped = 0

    def get_id(typ: str, name: str) -> int | None:
        row = conn.execute(
            "SELECT id FROM categories WHERE typ=? AND name=?",
            (typ, name),
        ).fetchone()
        return int(row["id"]) if row else None

    def upsert(typ: str, name: str, *, parent_id: int | None, is_fix: bool, is_rec: bool, day: int) -> int:
        nonlocal inserted, updated
        existing_id = get_id(typ, name)
        if has_parent:
            conn.execute(
                "INSERT INTO categories(typ,name,parent_id,is_fix,is_recurring,recurring_day) "
                "VALUES(?,?,?,?,?,?) "
                "ON CONFLICT(typ,name) DO UPDATE SET "
                "  parent_id=excluded.parent_id, "
                "  is_fix=excluded.is_fix, "
                "  is_recurring=excluded.is_recurring, "
                "  recurring_day=excluded.recurring_day",
                (typ, name, parent_id, int(is_fix), int(is_rec), int(day)),
            )
        else:
            conn.execute(
                "INSERT INTO categories(typ,name,is_fix,is_recurring,recurring_day) "
                "VALUES(?,?,?,?,?) "
                "ON CONFLICT(typ,name) DO UPDATE SET "
                "  is_fix=excluded.is_fix, "
                "  is_recurring=excluded.is_recurring, "
                "  recurring_day=excluded.recurring_day",
                (typ, name, int(is_fix), int(is_rec), int(day)),
            )

        if existing_id is None:
            inserted += 1
        else:
            updated += 1

        cid = get_id(typ, name)
        if cid is None:
            raise RuntimeError(f"Konnte Kategorie nicht anlegen: {typ} / {name}")
        return cid

    # Daten
    for r_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        typ_raw = row[c_typ].value if c_typ is not None else None
        path_raw = row[c_path].value if c_path is not None else None

        if typ_raw is None and path_raw is None:
            continue

        typ = _norm_typ(typ_raw)
        path_str = str(path_raw).strip() if path_raw is not None else ""
        if not typ or not path_str:
            skipped += 1
            continue

        if path_str.startswith("#"):
            continue

        if typ not in {"Einkommen", "Ausgaben", "Ersparnisse"}:
            warnings.append(f"Zeile {r_idx}: Unbekannter Typ '{typ_raw}' → übersprungen.")
            skipped += 1
            continue

        parts = _split_path(path_str)
        if not parts:
            skipped += 1
            continue

        is_fix = _as_bool(row[c_fix].value) if c_fix is not None else False
        is_rec = _as_bool(row[c_rec].value) if c_rec is not None else False
        day = _as_int_day(row[c_day].value) if c_day is not None else 1

        parent_id: int | None = None
        for i, name in enumerate(parts):
            leaf = i == (len(parts) - 1)
            # Elternknoten: Flags aus, Blatt: Flags aus Excel
            cid = upsert(
                typ,
                name,
                parent_id=parent_id,
                is_fix=is_fix if leaf else False,
                is_rec=is_rec if leaf else False,
                day=day if leaf else 1,
            )
            parent_id = cid

    conn.commit()
    return CategoryImportResult(inserted=inserted, updated=updated, skipped=skipped, warnings=warnings)
