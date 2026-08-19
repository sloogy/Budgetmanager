"""Sicherheitsgrenzen für vom Benutzer geladene Excel-Dateien.

XLSX/XLSM sind ZIP-Container mit XML-Dateien. Vor openpyxl werden deshalb
Dateigrösse, Memberzahl, entpackte Grösse, Einzelmember, Kompressionsrate,
Pfad-Traversal, Symlinks und DTD/ENTITY-Deklarationen geprüft.
"""

from __future__ import annotations

import stat
import zipfile
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

MAX_EXCEL_ARCHIVE_BYTES = 64 * 1024 * 1024
MAX_EXCEL_MEMBERS = 4_096
MAX_EXCEL_MEMBER_BYTES = 64 * 1024 * 1024
MAX_EXCEL_UNCOMPRESSED_BYTES = 256 * 1024 * 1024
MAX_EXCEL_COMPRESSION_RATIO = 200
MAX_XML_PROLOG_SCAN_BYTES = 128 * 1024
ALLOWED_SUFFIXES = frozenset({".xlsx", ".xlsm", ".xltx", ".xltm"})


class UnsafeExcelFileError(ValueError):
    """Excel-Datei verletzt eine Sicherheitsgrenze."""


def _is_symlink(info: zipfile.ZipInfo) -> bool:
    mode = (info.external_attr >> 16) & 0xFFFF
    return stat.S_ISLNK(mode)


def validate_excel_archive(path: Path) -> Path:
    """Validiert einen Office-Open-XML-Container fail-closed."""
    file_path = Path(path).expanduser().resolve()
    if not file_path.is_file():
        raise FileNotFoundError(str(file_path))
    if file_path.suffix.lower() not in ALLOWED_SUFFIXES:
        raise UnsafeExcelFileError("Nur XLSX/XLSM-Dateien sind erlaubt")
    archive_size = file_path.stat().st_size
    if archive_size <= 0 or archive_size > MAX_EXCEL_ARCHIVE_BYTES:
        raise UnsafeExcelFileError(
            f"Excel-Datei ist leer oder grösser als {MAX_EXCEL_ARCHIVE_BYTES // (1024 * 1024)} MB"
        )
    if not zipfile.is_zipfile(file_path):
        raise UnsafeExcelFileError("Datei ist kein gültiger XLSX/XLSM-Container")

    try:
        with zipfile.ZipFile(file_path, "r") as archive:
            members = archive.infolist()
            if not members or len(members) > MAX_EXCEL_MEMBERS:
                raise UnsafeExcelFileError(
                    "Excel-Datei ist leer oder enthält zu viele Einträge"
                )

            total_uncompressed = 0
            for info in members:
                name = info.filename
                member_path = Path(name)
                if member_path.is_absolute() or ".." in member_path.parts:
                    raise UnsafeExcelFileError(
                        f"Unsicherer Pfad im Excel-Archiv: {name}"
                    )
                if _is_symlink(info):
                    raise UnsafeExcelFileError(
                        f"Symlink im Excel-Archiv ist nicht erlaubt: {name}"
                    )
                if info.flag_bits & 0x1:
                    raise UnsafeExcelFileError(
                        "Passwortgeschützte Excel-Archive werden nicht importiert"
                    )
                if info.file_size > MAX_EXCEL_MEMBER_BYTES:
                    raise UnsafeExcelFileError(
                        f"Eintrag im Excel-Archiv ist zu gross: {name}"
                    )
                total_uncompressed += int(info.file_size)
                if total_uncompressed > MAX_EXCEL_UNCOMPRESSED_BYTES:
                    raise UnsafeExcelFileError(
                        "Excel-Datei ist entpackt unplausibel gross"
                    )
                if (
                    info.compress_size > 0
                    and info.file_size / info.compress_size
                    > MAX_EXCEL_COMPRESSION_RATIO
                ):
                    raise UnsafeExcelFileError(
                        f"Auffällige Kompressionsrate im Excel-Archiv: {name}"
                    )

                if name.lower().endswith(".xml") and info.file_size:
                    with archive.open(info, "r") as stream:
                        head = stream.read(MAX_XML_PROLOG_SCAN_BYTES).upper()
                    if b"<!DOCTYPE" in head or b"<!ENTITY" in head:
                        raise UnsafeExcelFileError(
                            f"DTD/ENTITY-Deklaration in Excel-XML ist nicht erlaubt: {name}"
                        )
    except zipfile.BadZipFile as exc:
        raise UnsafeExcelFileError(f"Ungültiger Excel-Container: {exc}") from exc

    return file_path


def load_workbook_safely(path: Path, **kwargs: Any):
    """Validiert und lädt eine Arbeitsmappe mit openpyxl.

    ``defusedxml`` ist Laufzeitabhängigkeit und wird von openpyxl automatisch
    für XML-Parser verwendet. Die vorgeschalteten ZIP-Grenzen verhindern
    zusätzlich Ressourcenerschöpfung vor dem XML-Parsing.
    """
    file_path = validate_excel_archive(path)
    return load_workbook(file_path, **kwargs)
