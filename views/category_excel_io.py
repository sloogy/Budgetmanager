from __future__ import annotations

import csv
from dataclasses import dataclass
from pathlib import Path
import re
import sqlite3
from typing import Any

from openpyxl import Workbook, load_workbook


import logging
from utils.i18n import tr, trf
from model.crypto import suspend_after_commit_autosave

logger = logging.getLogger(__name__)

TYP_ALIASES = {
    "einnahmen": tr("kpi.income"),
    "income": tr("kpi.income"),
    "einkommen": tr("kpi.income"),
    "lohn": tr("kpi.income"),
    "ausgaben": tr("kpi.expenses"),
    "expenses": tr("kpi.expenses"),
    "ersparnisse": tr("typ.Ersparnisse"),
    "sparen": tr("typ.Ersparnisse"),
    "savings": tr("typ.Ersparnisse"),
}


def _norm_typ(value: Any) -> str | None:
    if value is None:
        return None
    s = str(value).strip()
    if not s:
        return None
    key = s.strip().lower()
    return TYP_ALIASES.get(key, s)


def _as_bool(value: Any) -> bool:
    if value is None:
        return False
    if isinstance(value, bool):
        return value
    s = str(value).strip().lower()
    return s in {"1", "true", "ja", "j", "yes", "y", "x", "✓", "ok"}


def _as_int_day(value: Any, default: int = 1) -> int:
    try:
        d = int(value)
    except Exception:
        d = int(default)
    if d < 1:
        return 1
    if d > 31:
        return 31
    return d


def _split_path(path: str) -> list[str]:
    # Unterstützt: "Wohnen › Miete", "Wohnen > Miete", "Wohnen / Miete"
    parts = re.split(r"\s*(?:›|»|>|/|\\)\s*", str(path).strip())
    return [p.strip() for p in parts if p and str(p).strip()]


def _split_subpath_csv(value: str) -> list[str]:
    """Splitter für die CSV-Unterkategorie-Spalte.

    Trennt NUR am expliziten Hierarchie-Marker (›/»/>), damit Namen mit Schrägstrich
    wie "Miete/Hypothek" oder "ÖV (Abo/Billette)" als EIN Blatt erhalten bleiben.
    Eine tiefere Verschachtelung (Fallback) wird als "Kind › Enkel" ausgedrückt.
    """
    parts = re.split(r"\s*(?:›|»|>)\s*", str(value).strip())
    return [p.strip() for p in parts if p and str(p).strip()]


@dataclass
class CategoryImportResult:
    inserted: int
    updated: int
    skipped: int
    warnings: list[str]


# ---------------------------------------------------------------------------
# Geteilte Kern-Routine (xlsx UND csv nutzen sie — keine Doppel-Logik)
# ---------------------------------------------------------------------------


def _has_parent_col(conn: sqlite3.Connection) -> bool:
    cols = {r[1] for r in conn.execute("PRAGMA table_info(categories)").fetchall()}
    return "parent_id" in cols


def _get_cat_id(conn: sqlite3.Connection, typ: str, name: str) -> int | None:
    row = conn.execute(
        "SELECT id FROM categories WHERE typ=? AND name=?", (typ, name)
    ).fetchone()
    if not row:
        return None
    return int(row["id"]) if hasattr(row, "keys") else int(row[0])


def _ensure_category(
    conn: sqlite3.Connection,
    typ: str,
    name: str,
    *,
    parent_id: int | None,
    is_fix: bool,
    is_rec: bool,
    day: int,
    has_parent: bool,
    set_flags: bool,
) -> tuple[int, bool]:
    """Legt eine Kategorie an oder aktualisiert sie. Gibt (id, war_neu) zurück.

    set_flags=True  (Blattknoten): Flags/Tag/parent_id werden gesetzt bzw. überschrieben.
    set_flags=False (Elternknoten): nur anlegen falls fehlend — vorhandene Flags
                    eines Eltern-Eintrags (eigene Zeile) bleiben unangetastet.
    """
    existing = _get_cat_id(conn, typ, name)

    if set_flags:
        if has_parent:
            conn.execute(
                "INSERT INTO categories(typ,name,parent_id,is_fix,is_recurring,recurring_day) "
                "VALUES(?,?,?,?,?,?) "
                "ON CONFLICT(typ,name) DO UPDATE SET "
                "  parent_id=excluded.parent_id, is_fix=excluded.is_fix, "
                "  is_recurring=excluded.is_recurring, recurring_day=excluded.recurring_day",
                (typ, name, parent_id, int(is_fix), int(is_rec), int(day)),
            )
        else:
            conn.execute(
                "INSERT INTO categories(typ,name,is_fix,is_recurring,recurring_day) "
                "VALUES(?,?,?,?,?) "
                "ON CONFLICT(typ,name) DO UPDATE SET "
                "  is_fix=excluded.is_fix, is_recurring=excluded.is_recurring, "
                "  recurring_day=excluded.recurring_day",
                (typ, name, int(is_fix), int(is_rec), int(day)),
            )
    else:
        # Elternknoten: nur anlegen, vorhandene Werte NICHT überschreiben.
        if has_parent:
            conn.execute(
                "INSERT OR IGNORE INTO categories(typ,name,parent_id,is_fix,is_recurring,recurring_day) "
                "VALUES(?,?,?,?,?,?)",
                (typ, name, parent_id, 0, 0, 1),
            )
        else:
            conn.execute(
                "INSERT OR IGNORE INTO categories(typ,name,is_fix,is_recurring,recurring_day) "
                "VALUES(?,?,?,?,?)",
                (typ, name, 0, 0, 1),
            )

    cid = _get_cat_id(conn, typ, name)
    if cid is None:
        raise RuntimeError(
            trf("lbl.konnte_kategorie_nicht_anlegen", typ=typ, name=name)
        )
    return cid, (existing is None)


def _apply_path(
    conn: sqlite3.Connection,
    typ: str,
    parts: list[str],
    *,
    is_fix: bool,
    is_rec: bool,
    day: int,
    has_parent: bool,
) -> tuple[int, int]:
    """Legt eine Kategorie-Kette (Pfad) an: Eltern zuerst, dann Blatt mit Flags.

    Gibt (neu_eingefügt, aktualisiert) zurück.
    """
    inserted = 0
    updated = 0
    parent_id: int | None = None
    for i, name in enumerate(parts):
        leaf = i == (len(parts) - 1)
        cid, was_new = _ensure_category(
            conn,
            typ,
            name,
            parent_id=parent_id,
            is_fix=is_fix if leaf else False,
            is_rec=is_rec if leaf else False,
            day=day if leaf else 1,
            has_parent=has_parent,
            set_flags=leaf,
        )
        if was_new:
            inserted += 1
        elif leaf:
            updated += 1
        parent_id = cid
    return inserted, updated


def export_category_template_xlsx(out_path: Path) -> Path:
    """Erstellt eine einfache Excel-Vorlage für Kategorien."""
    out_path = Path(out_path)
    if out_path.suffix.lower() != ".xlsx":
        out_path = out_path.with_suffix(".xlsx")

    wb = Workbook()

    ws = wb.active
    ws.title = tr("tab.categories")

    headers = [
        "Typ",
        "Pfad",
        "Fix (0/1)",
        "Wiederkehrend (0/1)",
        "Tag (1-31)",
    ]
    ws.append(headers)

    # Beispiele (kann der User löschen)
    ws.append([tr("kpi.expenses"), "Wohnen › Miete", 1, 1, 1])
    ws.append([tr("kpi.expenses"), tr("lbl.gesundheit_krankenkasse_praemie"), 1, 1, 1])
    ws.append([tr("kpi.income"), "Lohn", 0, 1, 25])
    ws.append([tr("typ.Ersparnisse"), "Notgroschen", 0, 1, 1])

    # Spaltenbreite
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 44
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 10

    # Info-Sheet
    info = wb.create_sheet("Info")
    info["A1"] = "Budgetmanager – Kategorien-Import"
    info["A3"] = "Spalten:"
    info["A4"] = "Typ: Einkommen / Ausgaben / Ersparnisse (Synonyme: Einnahmen, Sparen)"
    info["A5"] = "Pfad: z.B. Gesundheit › Krankenkasse › Prämie"
    info["A6"] = "Fix: 1 = Fixkosten (⭐)"
    info["A7"] = "Wiederkehrend: 1 = wiederkehrend (∞)"
    info["A8"] = "Tag: Fälligkeitstag (1–31)"
    info["A10"] = (
        "Hinweis: Du kannst die Beispielzeilen löschen und deine Struktur eintragen."
    )

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return out_path


def import_categories_from_xlsx(
    conn: sqlite3.Connection, xlsx_path: Path
) -> CategoryImportResult:
    """Importiert Kategorien (inkl. Baum-Pfad) aus einer Excel-Datei."""
    xlsx_path = Path(xlsx_path)
    if not xlsx_path.exists():
        raise FileNotFoundError(str(xlsx_path))

    wb = load_workbook(xlsx_path, data_only=True)
    ws = (
        wb[tr("tab.categories")] if tr("tab.categories") in wb.sheetnames else wb.active
    )

    # Header lesen
    header_row = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    header_map: dict[str, int] = {}
    for idx, h in enumerate(header_row):
        if h is None:
            continue
        key = str(h).strip().lower()
        header_map[key] = idx

    def col(*names: str) -> int | None:
        for n in names:
            if n in header_map:
                return header_map[n]
        return None

    c_typ = col("typ")
    c_path = col("pfad", "path")
    c_fix = col("fix (0/1)", "fix", "fixkosten")
    c_rec = col("wiederkehrend (0/1)", "wiederkehrend", "recurring")
    c_day = col("tag (1-31)", "tag", "day")

    if c_typ is None or c_path is None:
        raise ValueError(
            "Excel-Header muss mindestens 'Typ' und 'Pfad' enthalten (Sheet: Kategorien)."
        )

    # Prüfen ob parent_id Spalte existiert
    cols = {r[1] for r in conn.execute("PRAGMA table_info(categories)").fetchall()}
    has_parent = "parent_id" in cols

    warnings: list[str] = []
    inserted = 0
    updated = 0
    skipped = 0

    # Daten
    with suspend_after_commit_autosave(conn):
        for r_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
            typ_raw = row[c_typ].value if c_typ is not None else None
            path_raw = row[c_path].value if c_path is not None else None

            if typ_raw is None and path_raw is None:
                continue

            typ = _norm_typ(typ_raw)
            path_str = str(path_raw).strip() if path_raw is not None else ""
            if not typ or not path_str:
                skipped += 1
                continue

            if path_str.startswith("#"):
                continue

            if typ not in {tr("kpi.income"), tr("kpi.expenses"), tr("typ.Ersparnisse")}:
                warnings.append(
                    trf("msg.excel_unbekannter_typ", r_idx=r_idx, typ_raw=typ_raw)
                )
                skipped += 1
                continue

            parts = _split_path(path_str)
            if not parts:
                skipped += 1
                continue

            is_fix = _as_bool(row[c_fix].value) if c_fix is not None else False
            is_rec = _as_bool(row[c_rec].value) if c_rec is not None else False
            day = _as_int_day(row[c_day].value) if c_day is not None else 1

            ins, upd = _apply_path(
                conn,
                typ,
                parts,
                is_fix=is_fix,
                is_rec=is_rec,
                day=day,
                has_parent=has_parent,
            )
            inserted += ins
            updated += upd

        conn.commit()
    return CategoryImportResult(
        inserted=inserted, updated=updated, skipped=skipped, warnings=warnings
    )


# ---------------------------------------------------------------------------
# CSV: Spaltenlösung (Hauptkategorie + Unterkategorie statt ›-Pfad)
# ---------------------------------------------------------------------------

CSV_HEADERS = [
    "Typ",
    "Hauptkategorie",
    "Unterkategorie",
    "Fix (0/1)",
    "Wiederkehrend (0/1)",
    "Tag (1-31)",
]


def _iter_descendants(node: dict, prefix: list[str]):
    """Liefert (Category, unterkategorie_label) für alle Nachfahren von node.

    Bei genau 2 Ebenen ist label = Kindname (kein ›). Tiefere Ebenen werden als
    "Kind › Enkel" zusammengefasst (Fallback, damit nichts verloren geht).
    """
    for ch in node.get("children", []):
        c = ch["cat"]
        label = " › ".join(prefix + [c.name])
        yield c, label
        yield from _iter_descendants(ch, prefix + [c.name])


def export_categories_csv(conn: sqlite3.Connection, out_path: Path) -> Path:
    """Exportiert alle Kategorien als CSV mit getrennten Spalten für Haupt-/Unterkategorie.

    Format (UTF-8 mit BOM, damit Excel Umlaute korrekt zeigt):
        Typ, Hauptkategorie, Unterkategorie, Fix (0/1), Wiederkehrend (0/1), Tag (1-31)

    - Top-Level-Kategorie: Unterkategorie-Spalte leer.
    - Unterkategorie:      Hauptkategorie = Name der Elternkategorie.
    Die Datei lässt sich manuell in Excel/LibreOffice/Editor bearbeiten und mit
    import_categories_from_csv() wieder einspielen.
    """
    out_path = Path(out_path)
    if out_path.suffix.lower() != ".csv":
        out_path = out_path.with_suffix(".csv")

    from model.category_model import CategoryModel
    from model.typ_constants import TYP_INCOME, TYP_EXPENSES, TYP_SAVINGS

    cm = CategoryModel(conn)
    rows: list[list] = []
    for typ in (TYP_INCOME, TYP_EXPENSES, TYP_SAVINGS):
        for root in cm.build_tree(cm.list(typ)):
            rc = root["cat"]
            rows.append(
                [
                    typ,
                    rc.name,
                    "",
                    int(rc.is_fix),
                    int(rc.is_recurring),
                    int(rc.recurring_day),
                ]
            )
            for c, label in _iter_descendants(root, []):
                rows.append(
                    [
                        typ,
                        rc.name,
                        label,
                        int(c.is_fix),
                        int(c.is_recurring),
                        int(c.recurring_day),
                    ]
                )

    out_path.parent.mkdir(parents=True, exist_ok=True)
    with open(out_path, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow(CSV_HEADERS)
        w.writerows(rows)
    logger.info("Kategorien als CSV exportiert: %d Zeilen → %s", len(rows), out_path)
    return out_path


def _sniff_delimiter(sample: str) -> str:
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t")
        return dialect.delimiter
    except Exception:
        # Heuristik: Semikolon ist im deutschen Excel verbreitet
        if sample.count(";") > sample.count(","):
            return ";"
        return ","


def import_categories_from_csv(
    conn: sqlite3.Connection, csv_path: Path
) -> CategoryImportResult:
    """Importiert Kategorien aus der Spalten-CSV (Haupt-/Unterkategorie getrennt).

    Tolerant: Trennzeichen (`,` `;` Tab) wird automatisch erkannt; Header-Namen
    in mehreren Schreibweisen akzeptiert; Eltern werden bei Bedarf automatisch
    angelegt. Idempotent dank ON CONFLICT-Upsert.
    """
    csv_path = Path(csv_path)
    if not csv_path.exists():
        raise FileNotFoundError(str(csv_path))

    raw = csv_path.read_text(encoding="utf-8-sig")
    delimiter = _sniff_delimiter(raw[:4096])
    reader = csv.reader(raw.splitlines(), delimiter=delimiter)

    rows = list(reader)
    if not rows:
        return CategoryImportResult(0, 0, 0, ["CSV ist leer."])

    header = [str(h or "").strip().lower() for h in rows[0]]

    def col(*names: str) -> int | None:
        for n in names:
            if n in header:
                return header.index(n)
        return None

    c_typ = col("typ", "type")
    c_haupt = col("hauptkategorie", "kategorie", "category", "haupt")
    c_unter = col("unterkategorie", "subkategorie", "subcategory", "unter", "sub")
    c_fix = col("fix (0/1)", "fix", "fixkosten")
    c_rec = col("wiederkehrend (0/1)", "wiederkehrend", "recurring")
    c_day = col("tag (1-31)", "tag", "day")

    if c_typ is None or c_haupt is None:
        raise ValueError(
            "CSV-Header muss mindestens 'Typ' und 'Hauptkategorie' enthalten."
        )

    has_parent = _has_parent_col(conn)
    valid_typen = {tr("kpi.income"), tr("kpi.expenses"), tr("typ.Ersparnisse")}

    warnings: list[str] = []
    inserted = updated = skipped = 0

    def cell(row: list, idx: int | None):
        if idx is None or idx >= len(row):
            return None
        return row[idx]

    with suspend_after_commit_autosave(conn):
        for r_idx, row in enumerate(rows[1:], start=2):
            if not row or all((str(x or "").strip() == "") for x in row):
                continue

            typ = _norm_typ(cell(row, c_typ))
            haupt = str(cell(row, c_haupt) or "").strip()
            unter = str(cell(row, c_unter) or "").strip()

            if not typ or not haupt or haupt.startswith("#"):
                skipped += 1
                continue
            if typ not in valid_typen:
                warnings.append(
                    trf("msg.excel_unbekannter_typ", r_idx=r_idx, typ_raw=typ)
                )
                skipped += 1
                continue

            # Pfad bauen: [Hauptkategorie] (+ Unterkategorie, ggf. tiefer via ›)
            parts = [haupt] + (_split_subpath_csv(unter) if unter else [])

            is_fix = _as_bool(cell(row, c_fix))
            is_rec = _as_bool(cell(row, c_rec))
            day = _as_int_day(cell(row, c_day))

            ins, upd = _apply_path(
                conn,
                typ,
                parts,
                is_fix=is_fix,
                is_rec=is_rec,
                day=day,
                has_parent=has_parent,
            )
            inserted += ins
            updated += upd

        conn.commit()
    logger.info(
        "Kategorien-CSV importiert: +%d neu, %d aktualisiert, %d übersprungen",
        inserted,
        updated,
        skipped,
    )
    return CategoryImportResult(
        inserted=inserted, updated=updated, skipped=skipped, warnings=warnings
    )


def export_category_template_csv(out_path: Path) -> Path:
    """Schreibt eine leere CSV-Vorlage mit Beispielzeilen (Spaltenlösung)."""
    out_path = Path(out_path)
    if out_path.suffix.lower() != ".csv":
        out_path = out_path.with_suffix(".csv")

    examples = [
        [tr("kpi.expenses"), "Wohnen", "", 0, 0, 1],
        [tr("kpi.expenses"), "Wohnen", "Miete/Hypothek", 1, 1, 1],
        [tr("kpi.expenses"), "Wohnen", "Nebenkosten", 1, 1, 1],
        [tr("kpi.income"), "Lohn (Netto)", "", 0, 1, 25],
        [tr("typ.Ersparnisse"), "Rücklagen", "Ferien", 0, 0, 1],
    ]
    out_path.parent.mkdir(parents=True, exist_ok=True)
    with open(out_path, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow(CSV_HEADERS)
        w.writerows(examples)
    return out_path
