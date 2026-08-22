from __future__ import annotations

import json
import sqlite3
import sys
import zipfile

import pytest

from model import diagnostics, restore_bundle
from tools import coverage_gate, performance_release_gate
from utils.ui_experience_mode import (
    MODE_ADVANCED,
    MODE_CUSTOM,
    MODE_SIMPLE,
    detect_mode,
    mode_payload,
)


class SettingsStub:
    def __init__(self, values: dict | None = None):
        self.values = dict(values or {})

    def get(self, key: str, default=None):
        return self.values.get(key, default)


def test_ui_experience_modes_are_complete_and_detectable() -> None:
    simple = mode_payload(MODE_SIMPLE)
    assert simple["show_categories_tab"] is False
    assert simple["tab_visibility"]["savings"] is False
    assert simple["cockpit_preset"] == "focus"
    assert detect_mode(SettingsStub(simple)) == MODE_SIMPLE

    advanced = mode_payload(MODE_ADVANCED)
    assert advanced["show_categories_tab"] is True
    assert all(advanced["tab_visibility"].values())
    assert advanced["cockpit_preset"] == "standard"
    assert detect_mode(SettingsStub(advanced)) == MODE_ADVANCED

    custom = dict(simple)
    custom["tab_visibility"] = dict(simple["tab_visibility"])
    custom["tab_visibility"]["savings"] = True
    assert detect_mode(SettingsStub(custom)) == MODE_CUSTOM


def test_database_health_contains_only_technical_metadata() -> None:
    conn = sqlite3.connect(":memory:")
    try:
        conn.execute("CREATE TABLE private_notes(id INTEGER PRIMARY KEY, secret TEXT)")
        conn.execute("INSERT INTO private_notes(secret) VALUES('do not export')")
        health = diagnostics.database_health(conn)
    finally:
        conn.close()
    assert health["available"] is True
    assert health["quick_check"] == "ok"
    assert health["application_table_count"] == 1
    payload = json.dumps(health, ensure_ascii=False)
    assert "do not export" not in payload
    assert "private_notes" not in payload


def test_diagnostic_zip_contains_database_health_without_database(
    tmp_path, monkeypatch
) -> None:
    monkeypatch.setenv("BUDGETMANAGER_APP_DIR", str(tmp_path))
    diagnostics.log_file_path().parent.mkdir(parents=True, exist_ok=True)
    diagnostics.log_file_path().write_text("test", encoding="utf-8")
    conn = sqlite3.connect(":memory:")
    conn.execute("CREATE TABLE sample(id INTEGER PRIMARY KEY)")
    try:
        report = diagnostics.create_diagnostic_report_zip(connection=conn)
    finally:
        conn.close()
    with zipfile.ZipFile(report) as zf:
        names = set(zf.namelist())
        assert "database_health.json" in names
        assert not any(name.endswith((".db", ".enc", ".bmr")) for name in names)
        health = json.loads(zf.read("database_health.json"))
        assert health["quick_check"] == "ok"


def test_atomic_copy_verified_keeps_old_destination_when_replace_fails(
    tmp_path, monkeypatch
) -> None:
    source = tmp_path / "source.enc"
    destination = tmp_path / "active.enc"
    source.write_bytes(b"new database")
    destination.write_bytes(b"old database")

    def fail_replace(_src, _dst):
        raise PermissionError("simulated read-only target")

    monkeypatch.setattr(restore_bundle.os, "replace", fail_replace)
    with pytest.raises(PermissionError):
        restore_bundle.atomic_copy_verified(source, destination)
    assert destination.read_bytes() == b"old database"
    assert not list(tmp_path.glob(".*.restore_tmp_*"))


def test_atomic_copy_verified_rejects_wrong_expected_hash(tmp_path) -> None:
    source = tmp_path / "source.db"
    destination = tmp_path / "active.db"
    source.write_bytes(b"new database")
    destination.write_bytes(b"old database")
    with pytest.raises(restore_bundle.BundleIntegrityError):
        restore_bundle.atomic_copy_verified(
            source, destination, expected_sha256="0" * 64
        )
    assert destination.read_bytes() == b"old database"


def test_coverage_gate_writes_unambiguous_summary(tmp_path, monkeypatch) -> None:
    coverage_json = tmp_path / "coverage_full.json"
    summary_json = tmp_path / "summary.json"
    coverage_json.write_text(
        json.dumps(
            {
                "totals": {"percent_covered": 55.0},
                "files": {
                    name: {"summary": {"percent_covered": minimum + 1}}
                    for name, minimum in coverage_gate.DEFAULT_MINIMUMS.items()
                },
            }
        ),
        encoding="utf-8",
    )
    monkeypatch.setattr(
        sys,
        "argv",
        [
            "coverage_gate.py",
            "--json",
            str(coverage_json),
            "--summary-json",
            str(summary_json),
        ],
    )
    assert coverage_gate.main() == 0
    summary = json.loads(summary_json.read_text(encoding="utf-8"))
    assert summary["passed"] is True
    assert summary["overall"]["actual"] == 55.0


def test_large_database_benchmark_smoke() -> None:
    result = performance_release_gate.run_benchmark(3_000)
    assert result["rows"] == 3_000
    assert result["seeded"]["categories"] == 100
    assert result["result_sizes"]["overview_full_year"] >= 12
    assert all(value >= 0 for value in result["timings_seconds"].values())


def test_visual_render_metrics_reject_blank_and_accept_real_content() -> None:
    pytest.importorskip("PySide6")
    from PySide6.QtGui import QColor, QImage, QPainter, QPixmap
    from PySide6.QtWidgets import QApplication

    from utils.release_self_test import _pixmap_render_metrics

    # Die Zuweisung haelt die QApplication am Leben - ohne sie kann sie
    # eingesammelt werden, waehrend die Messung noch laeuft.
    app = QApplication.instance() or QApplication(["test-visual-render"])  # noqa: F841
    blank = QImage(800, 500, QImage.Format_ARGB32)
    blank.fill(0xFF101010)
    with pytest.raises(RuntimeError, match="einfarbig"):
        _pixmap_render_metrics(QPixmap.fromImage(blank))

    low_diversity = QImage(800, 500, QImage.Format_ARGB32)
    painter = QPainter(low_diversity)
    for index, colour in enumerate(("#101010", "#202020", "#303030")):
        painter.fillRect(index * 267, 0, 267, 500, QColor(colour))
    painter.end()
    with pytest.raises(RuntimeError, match="einfarbig"):
        _pixmap_render_metrics(QPixmap.fromImage(low_diversity))

    image = QImage(800, 500, QImage.Format_ARGB32)
    for y in range(image.height()):
        for x in range(image.width()):
            image.setPixelColor(
                x,
                y,
                QColor(
                    (x * 17) % 256,
                    (y * 23) % 256,
                    (x + y) % 256,
                    255,
                ),
            )
    metrics = _pixmap_render_metrics(QPixmap.fromImage(image))
    assert metrics["sample_colours"] >= 8
    assert metrics["opaque_ratio"] == 1.0


def test_xlsx_report_export_uses_separate_sheets_and_numeric_values(tmp_path) -> None:
    from openpyxl import load_workbook

    from model.report_export import ReportSection, export_sections_xlsx

    out = tmp_path / "report.xlsx"
    export_sections_xlsx(
        [
            ReportSection("Tracking", ("Datum", "Betrag"), (("30.07.2026", 12.5),)),
            ReportSection("Budget", ("Jahr", "Betrag"), ((2026, 100.0),)),
        ],
        out,
    )
    workbook = load_workbook(out, data_only=True)
    assert workbook.sheetnames == ["Tracking", "Budget"]
    assert workbook["Tracking"]["B2"].value == 12.5
    assert workbook["Budget"].freeze_panes == "A2"


def test_pdf_html_report_is_black_white_readable_and_escaped() -> None:
    from model.report_export import ReportSection, sections_to_html

    html = sections_to_html(
        [
            ReportSection(
                "Ausgaben",
                ("Kategorie", "Betrag"),
                (("Miete <Test>", 123.4),),
            )
        ],
        title="Jahresbericht",
        subtitle="2026",
    )
    assert "Miete &lt;Test&gt;" in html
    assert "border-collapse" in html
    assert "background:#eee" in html
    assert "Jahresbericht" in html
