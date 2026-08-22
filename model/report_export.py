"""Berichtsexport für XLSX und PDF, getrennt von der GUI."""

from __future__ import annotations

import os
from collections.abc import Sequence
from dataclasses import dataclass
from html import escape
from pathlib import Path


@dataclass(frozen=True)
class ReportSection:
    title: str
    headers: tuple[str, ...]
    rows: tuple[tuple[object, ...], ...]


def _safe_sheet_title(title: str, used: set[str]) -> str:
    invalid = set("[]:*?/\\")
    base = "".join("_" if ch in invalid else ch for ch in str(title)).strip() or "Daten"
    base = base[:31]
    candidate = base
    suffix = 2
    while candidate.casefold() in used:
        marker = f" {suffix}"
        candidate = (base[: 31 - len(marker)] + marker).strip()
        suffix += 1
    used.add(candidate.casefold())
    return candidate


def export_sections_xlsx(
    sections: Sequence[ReportSection],
    out_path: Path,
    *,
    include_headers: bool = True,
) -> Path:
    """Schreibt jede Datengruppe in ein eigenes, lesbares Excel-Blatt."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font
    from openpyxl.utils import get_column_letter

    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    workbook.remove(workbook.active)
    used: set[str] = set()
    for section in sections:
        sheet = workbook.create_sheet(_safe_sheet_title(section.title, used))
        if include_headers:
            sheet.append(list(section.headers))
            for cell in sheet[1]:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center")
            sheet.freeze_panes = "A2"
            last_column = get_column_letter(max(1, len(section.headers)))
            sheet.auto_filter.ref = f"A1:{last_column}1"
        for row in section.rows:
            sheet.append(list(row))
        for column in range(1, max(1, len(section.headers)) + 1):
            values = [
                str(sheet.cell(row=row, column=column).value or "")
                for row in range(1, min(sheet.max_row, 300) + 1)
            ]
            longest = max((len(value) for value in values), default=0)
            width = min(48, max(10, longest + 2))
            sheet.column_dimensions[get_column_letter(column)].width = width
        sheet.sheet_view.showGridLines = False
    if not sections:
        workbook.create_sheet("Daten")
    tmp = out_path.with_name(f".{out_path.stem}.tmp{out_path.suffix}")
    try:
        workbook.save(tmp)
        # Windows' os.fsync() uses the CRT _commit() call, which rejects a
        # read-only descriptor with EBADF.  Keep the descriptor writable on
        # every platform; no bytes are modified here.
        with tmp.open("rb+") as handle:
            os.fsync(handle.fileno())
        os.replace(tmp, out_path)
    finally:
        workbook.close()
        tmp.unlink(missing_ok=True)
    return out_path


def _html_cell(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float):
        return escape(f"{value:,.2f}".replace(",", "’"))
    return escape(str(value))


def sections_to_html(
    sections: Sequence[ReportSection],
    *,
    title: str,
    subtitle: str = "",
    include_headers: bool = True,
    empty_label: str = "Keine Daten",
) -> str:
    """Erzeugt drucktaugliches, schwarzweiss-lesbares HTML."""
    parts = [
        "<html><head><meta charset='utf-8'><style>",
        "body{font-family:sans-serif;font-size:9pt;color:#111;}",
        "h1{font-size:18pt;margin:0 0 4mm 0;} h2{font-size:12pt;margin:6mm 0 2mm 0;}",
        ".subtitle{color:#444;margin-bottom:5mm;}",
        "table{border-collapse:collapse;width:100%;page-break-inside:auto;}",
        "tr{page-break-inside:avoid;} "
        "th,td{border:0.3mm solid #777;padding:1.2mm;vertical-align:top;}",
        "th{font-weight:bold;background:#eee;} .empty{font-style:italic;color:#555;}",
        "</style></head><body>",
        f"<h1>{escape(title)}</h1>",
    ]
    if subtitle:
        parts.append(f"<div class='subtitle'>{escape(subtitle)}</div>")
    for section in sections:
        parts.append(f"<h2>{escape(section.title)}</h2><table>")
        if include_headers:
            header_cells = "".join(
                f"<th>{escape(header)}</th>" for header in section.headers
            )
            parts.append(f"<thead><tr>{header_cells}</tr></thead>")
        parts.append("<tbody>")
        if section.rows:
            for row in section.rows:
                cells = "".join(f"<td>{_html_cell(value)}</td>" for value in row)
                parts.append(f"<tr>{cells}</tr>")
        else:
            span = max(1, len(section.headers))
            parts.append(
                f"<tr><td class='empty' colspan='{span}'>"
                f"{escape(empty_label)}</td></tr>"
            )
        parts.append("</tbody></table>")
    parts.append("</body></html>")
    return "".join(parts)


def export_sections_pdf(
    sections: Sequence[ReportSection],
    out_path: Path,
    *,
    title: str,
    subtitle: str = "",
    include_headers: bool = True,
    empty_label: str = "Keine Daten",
) -> Path:
    """Schreibt einen paginierten PDF-Bericht über Qt ohne Zusatzbibliothek."""
    from PySide6.QtCore import QMarginsF
    from PySide6.QtGui import QPageLayout, QPageSize, QPdfWriter, QTextDocument

    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    tmp = out_path.with_name(f".{out_path.stem}.tmp{out_path.suffix}")
    try:
        writer = QPdfWriter(str(tmp))
        writer.setTitle(title)
        writer.setCreator("Budgetmanager")
        writer.setPageSize(QPageSize(QPageSize.PageSizeId.A4))
        writer.setPageMargins(QMarginsF(12, 12, 12, 12), QPageLayout.Unit.Millimeter)
        writer.setResolution(144)
        document = QTextDocument()
        document.setHtml(
            sections_to_html(
                sections,
                title=title,
                subtitle=subtitle,
                include_headers=include_headers,
                empty_label=empty_label,
            )
        )
        document.print_(writer)
        del document
        del writer
        if not tmp.is_file() or tmp.stat().st_size < 1_000:
            raise OSError("PDF-Bericht wurde nicht vollständig erzeugt")
        tmp.replace(out_path)
    finally:
        tmp.unlink(missing_ok=True)
    return out_path
